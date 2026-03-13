#!/usr/bin/env python3
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill("solid", fgColor="0C2948")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INFO_FILL = PatternFill("solid", fgColor="F7FAFC")
SAMPLE_FILL = PatternFill("solid", fgColor="EAF6F2")
NOTE_FILL = PatternFill("solid", fgColor="FFF4D6")
THIN = Side(style="thin", color="D8E3EA")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_sheet(sheet, widths):
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = True


def header_row(sheet, headers):
    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = ALL_BORDER
        cell.alignment = CENTER
    sheet.row_dimensions[1].height = 24


def fill_rows(sheet, start_row, rows):
    for row_index, values in enumerate(rows, start=start_row):
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.border = ALL_BORDER
            cell.alignment = LEFT
            cell.fill = SAMPLE_FILL if row_index == start_row else INFO_FILL


def add_data_validation(sheet, sqref, formula):
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.prompt = "可从下拉中选择，也可手动输入"
    validation.error = "请从下拉项中选择，或检查输入内容"
    sheet.add_data_validation(validation)
    validation.add(sqref)


def main():
    repo_root = Path(__file__).resolve().parent.parent
    desktop = Path("/Users/bu/Desktop")
    output = repo_root / "中药材全国追踪-Excel模板.xlsx"
    desktop_output = desktop / "中药材全国追踪-Excel模板.xlsx"

    wb = Workbook()

    origin = wb.active
    origin.title = "产地行情"
    origin_headers = [
        "记录日期", "品种", "规格", "产区", "报价", "单位",
        "行情标签", "摘要", "来源网站", "来源链接", "备注"
    ]
    header_row(origin, origin_headers)
    fill_rows(
        origin,
        2,
        [
            [
                "2026-03-13", "丹参", "产地干货精品货", "山东平邑", 16.5, "元/kg",
                "走快", "山东省临沂市平邑县丹参产地近期走货快", "中药材天地网",
                "https://example.com/origin", "示例行，可直接覆盖"
            ]
        ],
    )
    for row in range(2, 302):
        for col in range(1, 12):
            cell = origin.cell(row=row, column=col)
            cell.border = ALL_BORDER
            if row > 2:
                cell.fill = INFO_FILL
            cell.alignment = CENTER if col in [1, 5, 6, 7, 9] else LEFT
        origin.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        origin.cell(row=row, column=5).number_format = "0.00"
    origin.auto_filter.ref = "A1:K301"
    style_sheet(origin, {"A": 14, "B": 14, "C": 20, "D": 18, "E": 10, "F": 10, "G": 12, "H": 38, "I": 14, "J": 36, "K": 22})

    market = wb.create_sheet("四大市场")
    market_headers = [
        "记录日期", "市场", "品种", "规格", "报价", "单位",
        "行情标签", "摘要", "来源网站", "来源链接", "备注"
    ]
    header_row(market, market_headers)
    fill_rows(
        market,
        2,
        [
            [
                "2026-03-13", "安国", "黄连", "鸡爪连", 190, "元/kg",
                "平稳", "安国市场黄连报价平稳，经营户按需走货", "药通网",
                "https://example.com/market", "示例行，可直接覆盖"
            ]
        ],
    )
    for row in range(2, 302):
        for col in range(1, 12):
            cell = market.cell(row=row, column=col)
            cell.border = ALL_BORDER
            if row > 2:
                cell.fill = INFO_FILL
            cell.alignment = CENTER if col in [1, 2, 5, 6, 7, 9] else LEFT
        market.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        market.cell(row=row, column=5).number_format = "0.00"
    market.auto_filter.ref = "A1:K301"
    style_sheet(market, {"A": 14, "B": 12, "C": 14, "D": 18, "E": 10, "F": 10, "G": 12, "H": 38, "I": 14, "J": 36, "K": 22})

    hotspot = wb.create_sheet("行业热点")
    hotspot_headers = [
        "记录日期", "类型", "标题", "关联品种", "关联地区/市场",
        "摘要", "来源网站", "来源链接", "备注"
    ]
    header_row(hotspot, hotspot_headers)
    fill_rows(
        hotspot,
        2,
        [
            [
                "2026-03-13", "政策", "某省发布中药材产业扶持新规", "当归",
                "甘肃", "示例：这里记录政策、新闻、专题快讯，不混在行情表里", "人工整理",
                "https://example.com/hotspot", "示例行，可直接覆盖"
            ]
        ],
    )
    for row in range(2, 202):
        for col in range(1, 10):
            cell = hotspot.cell(row=row, column=col)
            cell.border = ALL_BORDER
            if row > 2:
                cell.fill = INFO_FILL
            cell.alignment = CENTER if col in [1, 2, 7] else LEFT
        hotspot.cell(row=row, column=1).number_format = "yyyy-mm-dd"
    hotspot.auto_filter.ref = "A1:I201"
    style_sheet(hotspot, {"A": 14, "B": 12, "C": 28, "D": 14, "E": 18, "F": 40, "G": 14, "H": 36, "I": 22})

    options = wb.create_sheet("下拉选项")
    columns = {
        "A": ("单位", ["元/kg", "元/公斤", "元/斤", "元/克"]),
        "B": ("市场", ["亳州", "安国", "成都", "玉林"]),
        "C": ("行情标签", ["上扬", "坚挺", "平稳", "走快", "下滑", "产新", "关注"]),
        "D": ("来源网站", ["药通网", "中药材天地网", "人工整理", "其他"]),
        "E": ("热点类型", ["政策", "新闻", "标准", "专题", "天气", "其他"]),
    }
    for col, (title, items) in columns.items():
        options[f"{col}1"] = title
        for idx, item in enumerate(items, start=2):
            options[f"{col}{idx}"] = item
    options.sheet_state = "hidden"

    add_data_validation(origin, "F2:F301", "'下拉选项'!$A$2:$A$5")
    add_data_validation(origin, "G2:G301", "'下拉选项'!$C$2:$C$8")
    add_data_validation(origin, "I2:I301", "'下拉选项'!$D$2:$D$5")

    add_data_validation(market, "B2:B301", "'下拉选项'!$B$2:$B$5")
    add_data_validation(market, "F2:F301", "'下拉选项'!$A$2:$A$5")
    add_data_validation(market, "G2:G301", "'下拉选项'!$C$2:$C$8")
    add_data_validation(market, "I2:I301", "'下拉选项'!$D$2:$D$5")

    add_data_validation(hotspot, "B2:B201", "'下拉选项'!$E$2:$E$7")
    add_data_validation(hotspot, "G2:G201", "'下拉选项'!$D$2:$D$5")

    guide = wb.create_sheet("字段说明")
    guide_headers = ["工作表", "字段", "是否必填", "填写建议"]
    header_row(guide, guide_headers)
    guide_rows = [
        ["产地行情", "记录日期", "是", "填当天日期"],
        ["产地行情", "品种/规格/产区", "是", "尽量统一标准叫法，不要写成整句话"],
        ["产地行情", "报价/单位", "是", "只填数字，单位优先统一成元/kg"],
        ["产地行情", "行情标签", "否", "上扬、坚挺、平稳、走快、下滑、产新、关注"],
        ["四大市场", "市场", "是", "只填亳州、安国、成都、玉林"],
        ["四大市场", "摘要", "是", "记录当天市场里的关键信息"],
        ["行业热点", "类型", "是", "政策、新闻、标准、专题、天气等"],
        ["行业热点", "标题/摘要", "是", "不要混到行情表里，单独维护更清晰"],
    ]
    fill_rows(guide, 2, guide_rows)
    for row in range(2, len(guide_rows) + 2):
        for col in range(1, 5):
            guide.cell(row=row, column=col).fill = NOTE_FILL if guide.cell(row=row, column=3).value == "是" else INFO_FILL
    style_sheet(guide, {"A": 14, "B": 18, "C": 10, "D": 42})

    usage = wb.create_sheet("使用说明")
    usage.merge_cells("A1:E1")
    usage["A1"] = "中药材全国追踪 三表录入模板"
    usage["A1"].fill = HEADER_FILL
    usage["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    usage["A1"].alignment = CENTER
    notes = [
        "1. 产地行情放产区报价；四大市场只放亳州、安国、成都、玉林；行业热点单独放政策和新闻。",
        "2. 不再按股市方式强制维护涨跌幅，重点记录地点、报价、标签、摘要和来源。",
        "3. 规格尽量写成标准规格，例如“统货”“鸡爪连”“120头”，不要写成长句。",
        "4. 行业热点建议每天补 3-5 条，哪怕先只有标题和摘要，也比没有强。",
        "5. 网站后面会优先读取这三张工作表，所以新数据尽量按这套格式继续整理。",
    ]
    for idx, note in enumerate(notes, start=3):
        usage[f"A{idx}"] = note
    usage.column_dimensions["A"].width = 90

    wb.save(output)
    shutil.copy2(output, desktop_output)
    print(output)
    print(desktop_output)


if __name__ == "__main__":
    main()
