#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


FOUR_MARKETS = ["亳州", "安国", "成都", "玉林"]
HOTSPOT_PLACEHOLDERS = [
    {
        "name": "行业政策",
        "desc": "待补充药政、标准、追溯、产新及天气等政策信息。",
        "value": "待录入",
    },
    {
        "name": "行业新闻",
        "desc": "待补充重点会议、企业动态、市场联动与区域资讯。",
        "value": "待录入",
    },
    {
        "name": "专题快讯",
        "desc": "待补充与重点品种相关的专题整理和深度观察。",
        "value": "待录入",
    },
]


def text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def date_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    raw = text(value)
    return raw[:10]


def number_text(value, unit: str) -> str:
    if value in ("", None):
        return "待补充"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return text(value)
    number_str = f"{number:.2f}".rstrip("0").rstrip(".")
    suffix = unit or "元/kg"
    return f"{number_str} {suffix}"


def shorten(value: str, limit: int = 40) -> str:
    raw = text(value)
    if len(raw) <= limit:
        return raw
    return f"{raw[:limit].rstrip()}..."


def classify_tag(*parts: str) -> str:
    joined = " ".join(text(part) for part in parts)
    if any(keyword in joined for keyword in ["产新", "新货"]):
        return "产新"
    if any(keyword in joined for keyword in ["上浮", "上扬", "上涨", "走高", "抬升"]):
        return "上扬"
    if any(keyword in joined for keyword in ["坚挺", "偏强"]):
        return "坚挺"
    if any(keyword in joined for keyword in ["走货快", "畅快", "购销顺畅", "走动快", "走动良好", "走销", "走动尚可"]):
        return "走快"
    if any(keyword in joined for keyword in ["下滑", "下调", "回落", "疲软"]):
        return "下滑"
    if any(keyword in joined for keyword in ["平稳", "稳定", "正常", "不温不火", "价稳", "行情稳"]):
        return "平稳"
    return "关注"


def row_dict(row) -> dict[str, str]:
    headers = [text(cell) for cell in row[0]]
    values = row[1]
    return {header: values[index] for index, header in enumerate(headers)}


def collect_from_quote_sheet(ws):
    rows = []
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        data = {headers[index]: row[index] for index in range(len(headers))}
        rows.append(
            {
                "date": date_text(data.get("记录日期")),
                "herb": text(data.get("品种")),
                "spec": text(data.get("规格")),
                "unit": text(data.get("单位")) or "元/kg",
                "market": text(data.get("市场")),
                "location": text(data.get("产区")),
                "price": data.get("今日价"),
                "summary": text(data.get("备注")),
                "source": text(data.get("来源网站")),
                "url": text(data.get("来源链接")),
                "row_index": row_index,
            }
        )

    origins = [item for item in rows if item["market"] == "产地"]
    markets = [item for item in rows if item["market"] and item["market"] != "产地"]
    return origins, markets, []


def collect_from_multi_sheet(wb):
    def collect(sheet_name, mapping):
        sheet = wb[sheet_name]
        headers = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        records = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            data = {headers[index]: row[index] for index in range(len(headers))}
            record = {"row_index": row_index}
            for target, source in mapping.items():
                record[target] = data.get(source)
            records.append(record)
        return records

    origins = collect(
        "产地行情",
        {
            "date": "记录日期",
            "herb": "品种",
            "spec": "规格",
            "location": "产区",
            "price": "报价",
            "unit": "单位",
            "tag": "行情标签",
            "summary": "摘要",
            "source": "来源网站",
            "url": "来源链接",
            "note": "备注",
        },
    )
    for item in origins:
        item["date"] = date_text(item["date"])
        item["herb"] = text(item.get("herb"))
        item["spec"] = text(item.get("spec"))
        item["location"] = text(item.get("location"))
        item["unit"] = text(item.get("unit")) or "元/kg"
        item["summary"] = text(item.get("summary") or item.get("note"))
        item["source"] = text(item.get("source"))
        item["url"] = text(item.get("url"))
        item["market"] = "产地"

    markets = collect(
        "四大市场",
        {
            "date": "记录日期",
            "market": "市场",
            "herb": "品种",
            "spec": "规格",
            "price": "报价",
            "unit": "单位",
            "tag": "行情标签",
            "summary": "摘要",
            "source": "来源网站",
            "url": "来源链接",
            "note": "备注",
        },
    )
    for item in markets:
        item["date"] = date_text(item["date"])
        item["market"] = text(item.get("market"))
        item["herb"] = text(item.get("herb"))
        item["spec"] = text(item.get("spec"))
        item["unit"] = text(item.get("unit")) or "元/kg"
        item["summary"] = text(item.get("summary") or item.get("note"))
        item["source"] = text(item.get("source"))
        item["url"] = text(item.get("url"))
        item["location"] = item["market"]

    hotspots = collect(
        "行业热点",
        {
            "date": "记录日期",
            "kind": "类型",
            "title": "标题",
            "herb": "关联品种",
            "location": "关联地区/市场",
            "summary": "摘要",
            "source": "来源网站",
            "url": "来源链接",
            "note": "备注",
        },
    )
    for item in hotspots:
        item["date"] = date_text(item["date"])
        item["kind"] = text(item.get("kind")) or "热点"
        item["title"] = text(item.get("title"))
        item["herb"] = text(item.get("herb"))
        item["location"] = text(item.get("location"))
        item["summary"] = text(item.get("summary") or item.get("note"))
        item["source"] = text(item.get("source"))
        item["url"] = text(item.get("url"))

    return origins, markets, hotspots


def load_records(workbook_path: Path):
    wb = load_workbook(workbook_path, data_only=True)
    if {"产地行情", "四大市场", "行业热点"}.issubset(set(wb.sheetnames)):
        return collect_from_multi_sheet(wb)
    if "报价录入" in wb.sheetnames:
        return collect_from_quote_sheet(wb["报价录入"])
    raise SystemExit("Workbook does not contain a supported sheet structure.")


def build_dashboard(origins, markets, hotspots, source_path: Path):
    all_dates = sorted({item["date"] for item in origins + markets + hotspots if item.get("date")})
    latest_date = all_dates[-1] if all_dates else ""
    counts_by_date = Counter(item["date"] for item in origins + markets + hotspots if item.get("date"))
    herb_count = len({item["herb"] for item in origins + markets if item.get("herb")})
    source_count = len({item["source"] for item in origins + markets + hotspots if item.get("source")})

    focus_markets = [item for item in markets if item.get("market") in FOUR_MARKETS]
    focus_market_count = len(focus_markets)
    covered_markets = sorted({item["market"] for item in focus_markets})
    other_market_count = len(markets) - focus_market_count

    stats = [
        {"label": "近三日记录", "value": str(len(origins) + len(markets) + len(hotspots)), "sub": "当前已整理的数据条数", "tone": "green"},
        {"label": "覆盖品种", "value": str(herb_count), "sub": "来自产地与市场的药材品种", "tone": "gold"},
        {"label": "产地行情", "value": str(len(origins)), "sub": "当前这批数据以产地更新为主", "tone": "green"},
        {"label": "四大市场", "value": str(focus_market_count), "sub": f"已覆盖 {len(covered_markets)}/4，其他市场 {other_market_count} 条", "tone": "blue"},
        {"label": "行业热点", "value": str(len(hotspots)), "sub": "政策与新闻需要单独持续录入", "tone": "red" if not hotspots else "gold"},
    ]

    trend = {
        "labels": [date[5:] for date in all_dates],
        "values": [counts_by_date[date] for date in all_dates],
        "peak": max((counts_by_date[date] for date in all_dates), default=0),
    }

    latest_origins = [item for item in origins if item.get("date") == latest_date]
    latest_origins.sort(key=lambda item: item["row_index"], reverse=True)
    origin_signals = []
    seen_signal_keys = set()
    for item in latest_origins:
        signal_key = item.get("herb") or f"{item.get('location')}-{item.get('row_index')}"
        if signal_key in seen_signal_keys:
            continue
        seen_signal_keys.add(signal_key)
        origin_signals.append(
            {
                "name": item["herb"],
                "tag": classify_tag(item.get("spec"), item.get("summary")),
                "desc": f"{item['location']} · {number_text(item.get('price'), item.get('unit'))} · {shorten(item.get('summary'), 34)}",
            }
        )
        if len(origin_signals) == 5:
            break

    max_market_count = max((sum(1 for item in focus_markets if item["market"] == market) for market in FOUR_MARKETS), default=0)
    market_cards = []
    market_updates = []
    for market in FOUR_MARKETS:
        items = [item for item in focus_markets if item["market"] == market]
        latest_item = max(items, key=lambda item: (item["date"], item["row_index"]), default=None)
        count = len(items)
        score = round(count / max_market_count * 100) if max_market_count else 0
        if latest_item:
            desc = f"近3日 {count} 条 · 最新 {latest_item['herb']} {number_text(latest_item.get('price'), latest_item.get('unit'))}"
            market_updates.append(
                {
                    "name": market,
                    "desc": f"{latest_item['herb']} · {shorten(latest_item.get('summary'), 28)}",
                    "value": number_text(latest_item.get("price"), latest_item.get("unit")),
                }
            )
        else:
            desc = "当前还没有整理到该市场的真实记录"
            market_updates.append(
                {"name": market, "desc": "等待补充该市场的最新行情与报价", "value": "待补录"}
            )
        market_cards.append({"name": market, "desc": desc, "score": score, "metric": f"{count} 条"})

    origin_groups = defaultdict(list)
    for item in origins:
        origin_groups[item["location"]].append(item)
    origin_cards = []
    for location, items in sorted(origin_groups.items(), key=lambda entry: len(entry[1]), reverse=True)[:4]:
        herbs = []
        for item in items:
            herb = item.get("herb")
            if herb and herb not in herbs:
                herbs.append(herb)
        dates = sorted({item["date"] for item in items if item.get("date")})
        origin_cards.append(
            {
                "name": location,
                "meta": f"近3日更新 {len(items)} 条 · 重点品种 { ' / '.join(herbs[:3]) or '待补充' }",
                "hot": f"覆盖 {len(dates)} 天 · 最新 {dates[-1][5:] if dates else '--'}",
            }
        )

    hotspot_rows = []
    for item in hotspots:
        label = item.get("kind") or "热点"
        title = item.get("title") or item.get("herb") or "行业热点"
        hotspot_rows.append(
            {
                "name": title,
                "desc": shorten(item.get("summary"), 36) or "待补充摘要",
                "value": label,
                "date": item.get("date"),
                "row_index": item.get("row_index", 0),
            }
        )
    hotspot_rows.sort(key=lambda item: (item["date"], item["row_index"]), reverse=True)
    hotspot_rows = hotspot_rows[:5] or HOTSPOT_PLACEHOLDERS

    table_rows = []
    mixed_rows = []
    for item in origins:
        mixed_rows.append((item["date"], item["row_index"], "产地", item))
    for item in markets:
        mixed_rows.append((item["date"], item["row_index"], item.get("market") or "市场", item))
    mixed_rows.sort(key=lambda entry: (entry[0], entry[1]), reverse=True)
    for date, _, kind, item in mixed_rows[:10]:
        table_rows.append(
            {
                "date": date[5:] if date else "--",
                "type": kind,
                "herb": item.get("herb") or "待补充",
                "place": item.get("location") or item.get("market") or "--",
                "price": number_text(item.get("price"), item.get("unit")),
                "summary": shorten(item.get("summary"), 32) or "待补充摘要",
            }
        )

    note = (
        f"当前这批数据以产地行情为主，四大市场已覆盖 {len(covered_markets)}/4，"
        f"行业热点 {len(hotspots)} 条，后面建议把政策和新闻单独补进来。"
    )

    modules = [
        {"title": "产地行情录入", "desc": "先持续补产区报价、行情标签和摘要，这是首页最核心的数据层。", "priority": "P1"},
        {"title": "四大市场页", "desc": "围绕亳州、安国、成都、玉林拆出专页，做市场动态和重点品种。", "priority": "P1"},
        {"title": "行业热点流", "desc": "单独录政策、新闻、标准和专题快讯，不再用涨跌榜思路替代。", "priority": "P1"},
        {"title": "品种专题", "desc": "对丹参、黄连、当归等重点品种做专题页，沉淀长期价值。", "priority": "P2"},
    ]

    return {
        "meta": {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_file": source_path.name,
            "latest_date": latest_date,
            "available_dates": all_dates,
            "latest_count": counts_by_date[latest_date] if latest_date else 0,
            "day_span": len(all_dates),
            "covered_markets": covered_markets,
            "focus_market_count": focus_market_count,
            "source_count": source_count,
            "note": note,
        },
        "status": f"真实数据版 · 最新整理于 {latest_date}" if latest_date else "真实数据版 · 等待导入",
        "hero": {
            "eyebrow": "中药材产地与市场情报首页",
            "title": "先把产地行情、四大市场和行业热点，放到一个首页里。",
            "lead": "这版不再按股市的涨跌榜逻辑来做，而是围绕你现在真正能拿到的数据：产地信息、四大市场信息，以及后续要补的政策与新闻热点。",
            "source_strip": [
                f"来源网站：{source_count} 个",
                f"覆盖日期：{len(all_dates)} 天",
                f"当前数据源：{source_path.name}",
            ],
            "side_big": "产地情报已接入",
            "side_note": note,
        },
        "side_metrics": [
            {"label": "最新日期", "value": latest_date[5:] if latest_date else "--", "tone": "gold"},
            {"label": "覆盖天数", "value": str(len(all_dates)), "tone": "blue"},
            {"label": "市场覆盖", "value": f"{len(covered_markets)}/4", "tone": ""},
            {"label": "热点条数", "value": str(len(hotspots)), "tone": "green" if hotspots else "red"},
        ],
        "stats": stats,
        "trend": trend,
        "signals": origin_signals,
        "markets": market_cards,
        "origins": origin_cards,
        "market_updates": market_updates,
        "hotspots": hotspot_rows,
        "table_rows": table_rows,
        "modules": modules,
        "footer": {
            "left": "当前地址：yaocaiquanguo-site.pages.dev",
            "right": f"当前已接入真实 Excel 数据源：{source_path.name}",
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Build dashboard JSON from herb workbook.")
    parser.add_argument("source", help="Path to the source workbook.")
    parser.add_argument(
        "--output",
        default="public/data/dashboard.json",
        help="Output JSON path relative to the repository root.",
    )
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parent.parent
    source_path = Path(args.source).expanduser().resolve()
    output_path = (repo_root / args.output).resolve()

    origins, markets, hotspots = load_records(source_path)
    payload = build_dashboard(origins, markets, hotspots, source_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
